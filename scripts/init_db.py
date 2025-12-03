#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
GMP药指汇数据库初始化脚本
用于初始化数据库、分类和导入文档数据
"""

import sys
import os

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import csv
import math
from datetime import datetime, date
import argparse

# 添加项目路径到Python路径
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app import create_app, db
from app.models import Organization, Category, User, Document

# 尝试导入openpyxl用于读取Excel文件
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("警告: 未安装openpyxl，无法读取Excel价格列表文件")


def init_db(app):
    """初始化数据库表结构"""
    with app.app_context():
        # 创建所有表
        db.create_all()
        print("数据库表创建完成")


# -----------------------------
# 通用工具函数
# -----------------------------

def _strip(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_date_any(v):
    """将多种输入解析为 date 对象；无法解析返回 None"""
    if not v:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    # 常见格式尝试
    fmts = [
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
        "%Y-%m", "%Y/%m", "%Y.%m",
        "%Y%m%d",
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(s, fmt)
            # 若只到月份，规范化为月初
            if fmt in ("%Y-%m", "%Y/%m", "%Y.%m"):
                return date(dt.year, dt.month, 1)
            return dt.date()
        except ValueError:
            pass
    return None


def parse_datetime_any(v):
    """将多种输入解析为 datetime；无法解析返回 None"""
    if not v:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    if not s:
        return None
    fmts = [
        "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d", "%Y/%m/%d",
        "%Y.%m.%d %H:%M:%S", "%Y.%m.%d",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def load_price_map():
    """读取 data/price_list.xlsx -> {英文标题: 价格(int 向下取整)}"""
    price_dict = {}
    price_list_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'price_list.xlsx')
    if EXCEL_SUPPORT and os.path.exists(price_list_path):
        try:
            workbook = openpyxl.load_workbook(price_list_path)
            worksheet = workbook.active
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0]:
                    title = str(row[0]).strip()
                    val = row[1]
                    try:
                        if val is None or str(val).strip() == '':
                            continue
                        price = int(math.floor(float(val)))
                        price_dict[title] = price
                    except Exception:
                        continue
        except Exception as e:
            print(f"读取价格列表文件时出错: {e}")
    return price_dict


def get_or_create_org(org_name):
    if not org_name:
        return None
    org = Organization.query.filter_by(name=org_name).first()
    if not org:
        org = Organization(name=org_name)
        db.session.add(org)
        db.session.flush()
    return org


def get_or_create_category(org_id, cat_name):
    if not (org_id and cat_name):
        return None
    cat = Category.query.filter_by(name=cat_name, org_id=org_id).first()
    if not cat:
        cat = Category(name=cat_name, org_id=org_id)
        db.session.add(cat)
        db.session.flush()
    return cat

def init_organizations(app):
    """初始化组织数据"""
    with app.app_context():
        # 检查是否已有组织数据
        if Organization.query.first() is None:
            # 添加默认组织
            orgs = [
                Organization(name='ISPE'),
                Organization(name='PDA'),
                Organization(name='WHO'),
                Organization(name='FDA Guidance'),
                Organization(name='EMA'),
                Organization(name='PIC/S'),
                Organization(name='ICH'),
                Organization(name='APIC'),
            ]
            
            for org in orgs:
                db.session.add(org)
            
            db.session.commit()
            print(f"添加了 {len(orgs)} 个组织")
        else:
            print("组织数据已存在，跳过添加")

def init_categories(app):
    """初始化组织分类，根据CSV文件中的实际数据动态创建"""
    with app.app_context():
        # 用于存储已处理的分类，避免重复创建
        processed_categories = set()
        
        # 处理PDA分类
        pda = Organization.query.filter_by(name='PDA').first()
        if pda and os.path.exists(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'pda_documents.csv')):
            # 读取PDA CSV文件中的分类
            with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'pda_documents.csv'), 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    category_name = row.get('category')
                    if category_name and category_name.strip():
                        category_key = (pda.id, category_name.strip())
                        if category_key not in processed_categories:
                            existing_cat = Category.query.filter_by(name=category_name.strip(), org_id=pda.id).first()
                            if not existing_cat:
                                cat = Category(name=category_name.strip(), org_id=pda.id)
                                db.session.add(cat)
                                print(f"添加了PDA分类: {category_name.strip()}")
                            processed_categories.add(category_key)
        
        # 处理WHO分类
        who = Organization.query.filter_by(name='WHO').first()
        if who and os.path.exists(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'who_documents.csv')):
            # 读取WHO CSV文件中的分类
            with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'who_documents.csv'), 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    category_name = row.get('category')
                    if category_name and category_name.strip():
                        category_key = (who.id, category_name.strip())
                        if category_key not in processed_categories:
                            existing_cat = Category.query.filter_by(name=category_name.strip(), org_id=who.id).first()
                            if not existing_cat:
                                cat = Category(name=category_name.strip(), org_id=who.id)
                                db.session.add(cat)
                                print(f"添加了WHO分类: {category_name.strip()}")
                            processed_categories.add(category_key)
        
        # 处理ISPE分类
        ispe = Organization.query.filter_by(name='ISPE').first()
        if ispe and os.path.exists(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'ispe_documents.csv')):
            # 读取ISPE CSV文件中的分类
            with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'ispe_documents.csv'), 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    category_name = row.get('category')
                    if category_name and category_name.strip():
                        category_key = (ispe.id, category_name.strip())
                        if category_key not in processed_categories:
                            existing_cat = Category.query.filter_by(name=category_name.strip(), org_id=ispe.id).first()
                            if not existing_cat:
                                cat = Category(name=category_name.strip(), org_id=ispe.id)
                                db.session.add(cat)
                                print(f"添加了ISPE分类: {category_name.strip()}")
                            processed_categories.add(category_key)
        
        # 处理FDA Guidance分类
        fda = Organization.query.filter_by(name='FDA Guidance').first()
        if fda and os.path.exists(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'fda_guidance_documents.csv')):
            # 读取FDA Guidance CSV文件中的分类
            with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'fda_guidance_documents.csv'), 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    category_name = row.get('category')
                    if category_name and category_name.strip():
                        category_key = (fda.id, category_name.strip())
                        if category_key not in processed_categories:
                            existing_cat = Category.query.filter_by(name=category_name.strip(), org_id=fda.id).first()
                            if not existing_cat:
                                cat = Category(name=category_name.strip(), org_id=fda.id)
                                db.session.add(cat)
                                print(f"添加了FDA Guidance分类: {category_name.strip()}")
                            processed_categories.add(category_key)
        
        # 处理APIC分类
        apic = Organization.query.filter_by(name='APIC').first()
        if apic and os.path.exists(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'apic_documents.csv')):
            # 读取APIC CSV文件中的分类
            with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'apic_documents.csv'), 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    category_name = row.get('category')
                    if category_name and category_name.strip():
                        category_key = (apic.id, category_name.strip())
                        if category_key not in processed_categories:
                            existing_cat = Category.query.filter_by(name=category_name.strip(), org_id=apic.id).first()
                            if not existing_cat:
                                cat = Category(name=category_name.strip(), org_id=apic.id)
                                db.session.add(cat)
                                print(f"添加了APIC分类: {category_name.strip()}")
                            processed_categories.add(category_key)
        
        # 提交所有更改
        db.session.commit()
        print("所有分类初始化完成")

def init_admin_user(app):
    """初始化默认管理员账户"""
    with app.app_context():
        # 检查是否已有用户数据
        if User.query.first() is None:
            # 添加默认管理员账户
            admin_email = app.config.get('GMP_SEEKER_ADMIN') or 'admin@qq.com'
            admin = User(
                username='admin',
                email=admin_email,
                role='admin'
            )
            admin.set_password('password')  # 默认密码，建议在首次登录后修改
            db.session.add(admin)
            db.session.commit()
            print(f"添加了默认管理员账户: {admin_email}")
        else:
            print("用户数据已存在，跳过添加管理员账户")


def import_documents_from_csv(csv_file_path, app, org_name):
    """通用文档导入函数"""
    with app.app_context():
        org = Organization.query.filter_by(name=org_name).first()
        if not org:
            print(f"错误: 未找到{org_name}组织")
            return
        
        # 读取价格列表（如果存在）
        price_dict = load_price_map()

        # 读取已导出的文档链接信息（如果存在）
        # 文件名为 data/documents_export.xlsx
        excel_link_map = {}
        export_xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'documents_export.xlsx')
        if os.path.exists(export_xlsx_path):
            if not EXCEL_SUPPORT:
                print("警告: 检测到 documents_export.xlsx，但未安装openpyxl，无法读取补充链接数据")
            else:
                try:
                    wb = openpyxl.load_workbook(export_xlsx_path)
                    ws = wb.active
                    # 读取表头
                    headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    # 映射所需列索引
                    def col_idx(name):
                        try:
                            return headers.index(name)
                        except ValueError:
                            return None

                    idx_title = col_idx('英文标题')
                    idx_ori_file = col_idx('原版文档链接')
                    idx_chn_file = col_idx('中文版文档链接')
                    idx_ori_prev = col_idx('原版预览链接')
                    idx_chn_prev = col_idx('中文版预览链接')

                    if idx_title is None:
                        print("警告: documents_export.xlsx 缺少 '英文标题' 列，无法匹配标题")
                    else:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not row:
                                continue
                            title_val = row[idx_title] if idx_title is not None and idx_title < len(row) else None
                            if not title_val:
                                continue
                            title_key = str(title_val).strip()
                            if not title_key:
                                continue
                            def safe_get(i):
                                if i is None or i >= len(row):
                                    return None
                                v = row[i]
                                return str(v).strip() if v is not None else None
                            excel_link_map[title_key] = {
                                'original_file_url': safe_get(idx_ori_file),
                                'translation_file_url': safe_get(idx_chn_file),
                                'original_preview_url': safe_get(idx_ori_prev),
                                'translation_preview_url': safe_get(idx_chn_prev),
                            }
                except Exception as e:
                    print(f"读取 documents_export.xlsx 文件时出错: {e}")

        # 检查CSV文件是否存在
        if not os.path.exists(csv_file_path):
            print(f"警告: {org_name} CSV文件不存在 {csv_file_path}")
            return
        
        # 读取CSV文件
        with open(csv_file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            imported_count = 0
            skipped_count = 0
            
            for row in reader:
                # 检查文档是否已存在
                existing_doc = Document.query.filter_by(title=row['title'].strip()).first()
                if existing_doc:
                    skipped_count += 1
                    continue
                
                # 查找分类（可选）
                category = None
                if row.get('category') and row['category'].strip():
                    category = Category.query.filter_by(name=row['category'].strip(), org_id=org.id).first()
                
                # 解析日期
                publish_date = parse_date_any(row.get('publish_date', '').strip() if row.get('publish_date') else None)
                
                # 获取价格信息
                title = row['title'].strip()
                price = int(price_dict.get(title, 0))  # 默认价格为0

                # 从CSV读取文件与预览链接；缺失时尝试用Excel补全
                csv_original_file_url = (row.get('original_file_url') or '').strip()
                csv_translation_file_url = (row.get('translation_file_url') or '').strip()
                csv_original_preview_url = (row.get('original_preview_url') or '').strip()
                csv_translation_preview_url = (row.get('translation_preview_url') or '').strip()

                # 默认使用CSV中的值
                original_file_url = csv_original_file_url or None
                translation_file_url = csv_translation_file_url or None
                original_preview_url = csv_original_preview_url or None
                translation_preview_url = csv_translation_preview_url or None

                # 如果为空且Excel有对应标题，则使用Excel中的值
                link_row = excel_link_map.get(title)
                if link_row:
                    if not original_file_url:
                        v = (link_row.get('original_file_url') or '').strip() if link_row.get('original_file_url') else None
                        original_file_url = v or original_file_url
                    if not translation_file_url:
                        v = (link_row.get('translation_file_url') or '').strip() if link_row.get('translation_file_url') else None
                        translation_file_url = v or translation_file_url
                    if not original_preview_url:
                        v = (link_row.get('original_preview_url') or '').strip() if link_row.get('original_preview_url') else None
                        original_preview_url = v or original_preview_url
                    if not translation_preview_url:
                        v = (link_row.get('translation_preview_url') or '').strip() if link_row.get('translation_preview_url') else None
                        translation_preview_url = v or translation_preview_url
                
                # 创建文档对象（完全统一的字段处理）
                doc = Document(
                    title=title,
                    chinese_title=row.get('chinese_title', '').strip() or None,
                    summary=row.get('summary', '').strip() or None,
                    chinese_summary=row.get('chinese_summary', '').strip() or None,
                    org_id=org.id,
                    category_id=category.id if category else None,
                    cover_url=row.get('cover_url', '').strip() or None,
                    publish_date=publish_date,
                    source_url=row.get('source_url', '').strip() or None,
                    original_file_url=original_file_url,
                    translation_file_url=translation_file_url,
                    original_preview_url=original_preview_url,
                    translation_preview_url=translation_preview_url,
                    price=price
                )
                
                db.session.add(doc)
                imported_count += 1
                
                # 每100条记录提交一次
                if imported_count % 100 == 0:
                    db.session.commit()
            
            # 提交剩余记录
            db.session.commit()
            print(f"{org_name}文档导入完成: 成功导入 {imported_count} 条记录，跳过 {skipped_count} 条记录")


def import_documents_from_excel(excel_path, app, org_filter=None, upsert=False, dry_run=False):
    """从 Excel 文件导入文档（优先数据源）

    列要求（严格匹配中文列名，自动去空格）：
    ID, 组织, 分类, 英文标题, 中文标题, 概述, 中文概述, 封面链接, 出版日期,
    源链接, 原版文档链接, 中文版文档链接, 原版预览链接, 中文版预览链接, 价格, 创建时间, 更新时间
    """
    if not EXCEL_SUPPORT:
        print("错误: 未安装 openpyxl，无法从 Excel 导入")
        return

    if not os.path.exists(excel_path):
        print(f"错误: Excel 文件不存在: {excel_path}")
        return

    price_map = load_price_map()

    def norm_col(v):
        return (str(v).strip() if v is not None else '')

    with app.app_context():
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        except Exception as e:
            print(f"读取 Excel 失败: {e}")
            return

        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [norm_col(c.value) for c in header_cells]
        header_index = {h: i for i, h in enumerate(headers)}

        required_cols = [
            '组织', '分类', '英文标题', '中文标题', '概述', '中文概述', '封面链接', '出版日期',
            '源链接', '原版文档链接', '中文版文档链接', '原版预览链接', '中文版预览链接', '价格', '创建时间', '更新时间'
        ]
        # 仅“英文标题”为必需；其余缺失可为空
        if '英文标题' not in header_index:
            print("错误: Excel 缺少必需列 ‘英文标题’")
            return

        imported = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # 读取列值并规范化
            def get(col):
                idx = header_index.get(col)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            org_name = _strip(get('组织'))
            cat_name = _strip(get('分类'))
            title = _strip(get('英文标题'))
            if not title:
                skipped += 1
                continue

            if org_filter and org_name != org_filter:
                continue

            chinese_title = _strip(get('中文标题'))
            summary = _strip(get('概述'))
            chinese_summary = _strip(get('中文概述'))
            cover_url = _strip(get('封面链接'))
            publish_date = parse_date_any(get('出版日期'))
            source_url = _strip(get('源链接'))
            original_file_url = _strip(get('原版文档链接'))
            translation_file_url = _strip(get('中文版文档链接'))
            original_preview_url = _strip(get('原版预览链接'))
            translation_preview_url = _strip(get('中文版预览链接'))

            # 价格：优先 Excel，否则 price_list.xlsx，否则 0；向下取整
            price_val = get('价格')
            price = None
            try:
                if price_val is not None and str(price_val).strip() != '':
                    price = int(math.floor(float(price_val)))
            except Exception:
                price = None
            if price is None:
                price = int(price_map.get(title, 0))

            created_at = parse_datetime_any(get('创建时间'))
            updated_at = parse_datetime_any(get('更新时间'))

            # 组织与分类按需创建（dry-run 下只查询不创建）
            if dry_run:
                org = Organization.query.filter_by(name=org_name).first() if org_name else None
                category = Category.query.filter_by(name=cat_name, org_id=org.id if org else None).first() if (org and cat_name) else None
            else:
                org = get_or_create_org(org_name) if org_name else None
                category = get_or_create_category(org.id, cat_name) if (org and cat_name) else None

            existing = Document.query.filter_by(title=title).first()
            if existing:
                if not upsert:
                    skipped += 1
                    continue
                if dry_run:
                    updated += 1
                else:
                    # upsert：非空覆盖
                    if org:
                        existing.org_id = org.id
                    if category:
                        existing.category_id = category.id
                    if chinese_title is not None:
                        existing.chinese_title = chinese_title
                    if summary is not None:
                        existing.summary = summary
                    if chinese_summary is not None:
                        existing.chinese_summary = chinese_summary
                    if cover_url is not None:
                        existing.cover_url = cover_url
                    if publish_date is not None:
                        existing.publish_date = publish_date
                    if source_url is not None:
                        existing.source_url = source_url
                    if original_file_url is not None:
                        existing.original_file_url = original_file_url
                    if translation_file_url is not None:
                        existing.translation_file_url = translation_file_url
                    if original_preview_url is not None:
                        existing.original_preview_url = original_preview_url
                    if translation_preview_url is not None:
                        existing.translation_preview_url = translation_preview_url
                    if price is not None:
                        existing.price = int(price)
                    if created_at is not None:
                        existing.created_at = created_at
                    if updated_at is not None:
                        existing.updated_at = updated_at
                    updated += 1
            else:
                if dry_run:
                    imported += 1
                else:
                    doc = Document(
                        title=title,
                        chinese_title=chinese_title,
                        summary=summary,
                        chinese_summary=chinese_summary,
                        org_id=org.id if org else None,
                        category_id=category.id if category else None,
                        cover_url=cover_url,
                        publish_date=publish_date,
                        source_url=source_url,
                        original_file_url=original_file_url,
                        translation_file_url=translation_file_url,
                        original_preview_url=original_preview_url,
                        translation_preview_url=translation_preview_url,
                        price=int(price) if price is not None else 0,
                    )
                    if created_at is not None:
                        doc.created_at = created_at
                    if updated_at is not None:
                        doc.updated_at = updated_at
                    db.session.add(doc)
                    imported += 1

            # 分批提交
            total = imported + updated
            if not dry_run and total > 0 and total % 100 == 0:
                db.session.commit()

        if not dry_run:
            db.session.commit()
        print(f"Excel 导入完成: 新增 {imported}，更新 {updated}，跳过 {skipped}")

def import_pda_documents_from_csv(csv_file_path, app):
    """从PDA CSV文件导入文档数据"""
    return import_documents_from_csv(csv_file_path, app, 'PDA')

def import_who_documents_from_csv(csv_file_path, app):
    """从WHO CSV文件导入文档数据"""
    return import_documents_from_csv(csv_file_path, app, 'WHO')

def import_ispe_documents_from_csv(csv_file_path, app):
    """从ISPE CSV文件导入文档数据"""
    return import_documents_from_csv(csv_file_path, app, 'ISPE')

def import_fda_guidance_documents_from_csv(csv_file_path, app):
    """从FDA Guidance CSV文件导入文档数据"""
    return import_documents_from_csv(csv_file_path, app, 'FDA Guidance')

def import_apic_documents_from_csv(csv_file_path, app):
    """从APIC CSV文件导入文档数据"""
    return import_documents_from_csv(csv_file_path, app, 'APIC')

def import_all_documents(app):
    """导入所有文档"""
    print("开始导入PDA文档...")
    pda_csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'pda_documents.csv')
    if os.path.exists(pda_csv_path):
        import_pda_documents_from_csv(pda_csv_path, app)
    else:
        print(f"警告: PDA CSV文件不存在 {pda_csv_path}")
    
    print("开始导入WHO文档...")
    who_csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'who_documents.csv')
    if os.path.exists(who_csv_path):
        import_who_documents_from_csv(who_csv_path, app)
    else:
        print(f"警告: WHO CSV文件不存在 {who_csv_path}")
    
    print("开始导入ISPE文档...")
    ispe_csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'ispe_documents.csv')
    if os.path.exists(ispe_csv_path):
        import_ispe_documents_from_csv(ispe_csv_path, app)
    else:
        print(f"警告: ISPE CSV文件不存在 {ispe_csv_path}")
    
    print("开始导入FDA Guidance文档...")
    fda_csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'fda_guidance_documents.csv')
    if os.path.exists(fda_csv_path):
        import_fda_guidance_documents_from_csv(fda_csv_path, app)
    else:
        print(f"警告: FDA Guidance CSV文件不存在 {fda_csv_path}")
    
    print("开始导入APIC文档...")
    apic_csv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'apic_documents.csv')
    if os.path.exists(apic_csv_path):
        import_apic_documents_from_csv(apic_csv_path, app)
    else:
        print(f"警告: APIC CSV文件不存在 {apic_csv_path}")
    
    print("所有文档导入完成")


def import_all_documents_auto(app, source='auto', org_filter=None, upsert=False, dry_run=False):
    """根据来源选择导入（优先 Excel）"""
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')
    xlsx = os.path.join(data_dir, 'documents_export.xlsx')

    def can_excel():
        return EXCEL_SUPPORT and os.path.exists(xlsx)

    if source not in ('auto', 'excel', 'csv'):
        source = 'auto'

    if source == 'excel' or (source == 'auto' and can_excel()):
        if not can_excel():
            print("警告: 选择了 Excel 但不满足条件（缺文件或无 openpyxl），降级为 CSV")
        else:
            return import_documents_from_excel(xlsx, app, org_filter=org_filter, upsert=upsert, dry_run=dry_run)

    # CSV 路径导入（保持原逻辑）
    return import_all_documents(app)

def init_db_comprehensive(source='auto', org_filter=None, upsert=False, dry_run=False):
    """完整初始化数据库"""
    # 创建应用实例（遵循与 run.py 一致的配置选择）
    # 优先使用环境变量 FLASK_ENV 指定的配置，否则回落到默认配置
    app = create_app(os.getenv('FLASK_ENV') or 'default')
    
    print("开始初始化数据库...")
    init_db(app)
    
    print("开始初始化组织...")
    init_organizations(app)
    
    print("开始初始化分类...")
    init_categories(app)
    
    print("开始初始化管理员账户...")
    init_admin_user(app)
    
    print("开始导入文档...")
    import_all_documents_auto(app, source=source, org_filter=org_filter, upsert=upsert, dry_run=dry_run)
    
    print("数据库完整初始化完成")


def import_new_documents():
    """增量导入新文档（从单独的脚本导入）"""
    # 这个函数已被移到 import_new_documents.py 脚本中
    # 保持此函数以维持向后兼容性
    print("请使用 'python scripts/import_new_documents.py' 命令来增量导入新文档")


if __name__ == '__main__':
    # 向后兼容的增量参数
    if len(sys.argv) > 1 and sys.argv[1] == 'incremental':
        import_new_documents()
    else:
        parser = argparse.ArgumentParser(description='初始化数据库并导入数据')
        parser.add_argument('--source', default='auto', choices=['auto', 'excel', 'csv'], help='数据来源：auto(默认) | excel | csv')
        parser.add_argument('--org', dest='org_filter', default=None, help='仅导入指定组织（与 Excel/CSV 皆可配合）')
        parser.add_argument('--upsert', action='store_true', help='遇到已存在标题时更新（默认跳过）')
        parser.add_argument('--dry-run', action='store_true', help='仅统计与校验，不写入数据库')
        args = parser.parse_args()

        init_db_comprehensive(source=args.source, org_filter=args.org_filter, upsert=args.upsert, dry_run=args.dry_run)
